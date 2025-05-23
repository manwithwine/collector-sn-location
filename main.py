import os
from datetime import datetime
from netmiko import ConnectHandler
import openpyxl
from openpyxl.styles import Font
from dotenv import load_dotenv
from collector.collector_factory import CollectorFactory


def read_ip_addresses(filename):
    with open(filename, 'r') as file:
        return [line.strip() for line in file if line.strip()]


def create_excel_report(folder_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device Info"
    headers = ["Hostname", "S/N", "Location", "IP Address", "Vendor"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    return wb, ws


def determine_device_type(net_connect):
    output = net_connect.send_command("show version")
    if 'Huawei' in output:
        return 'Huawei'
    elif 'Cisco' in output:
        return 'Cisco'
    elif 'BCOM' in output:
        return 'B4COM'
    elif 'B4TECH' in output:
        return 'B4TECH'
    return None


def main():
    load_dotenv()
    ip_addresses = read_ip_addresses("ip")

    # Create output folder
    folder_name = f"results/cfg_{datetime.now().strftime('%Y%m%d_%H%M')}"
    os.makedirs(folder_name, exist_ok=True)

    wb, ws = create_excel_report(folder_name)
    row_counter = 2

    for ip_address in ip_addresses:
        print(f"Connecting to {ip_address}...")
        try:
            username = os.getenv('NET_USERNAME')
            password = os.getenv('NET_PASSWORD')

            with ConnectHandler(
                    device_type='autodetect',
                    host=ip_address,
                    username=username,
                    password=password,
                    global_delay_factor=2,
                    read_timeout_override=100,
                    timeout=30,
                    fast_cli=False
            ) as net_connect:
                print(f"Connected to {ip_address}")

                # Auto-detect and create collector
                collector = CollectorFactory.create_collector(net_connect)
                if not collector:
                    print(f"Unsupported device type for {ip_address}")
                    continue

                # Collect and process data
                result = collector.collect()

                # Save to Excel
                ws.cell(row=row_counter, column=1, value=result['hostname'])
                ws.cell(row=row_counter, column=2, value=result['serial'])
                ws.cell(row=row_counter, column=3, value=result['location'])
                ws.cell(row=row_counter, column=4, value=ip_address)
                ws.cell(row=row_counter, column=5, value=result['vendor'])
                row_counter += 1

                # Save raw output
                filename = f"{result['hostname']}.txt"
                with open(f"{folder_name}/{filename}", 'w') as f:
                    f.write(str(result))

                print(f"Saved data for {ip_address}")

        except Exception as e:
            print(f"Error with {ip_address}: {str(e)}")

    # Save Excel file
    excel_file = f"{folder_name}/device_inventory.xlsx"
    wb.save(excel_file)
    print(f"Report saved to {excel_file}")


if __name__ == "__main__":
    main()